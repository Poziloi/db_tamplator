from openpyxl import load_workbook
from flask_sqlalchemy import SQLAlchemy
from werkzeug.security import generate_password_hash

from models import User
from app import db, create_app

app = create_app()
app.app_context().push()

number_of_users = int(input('Введите количество пользователей, которых необходимо добавить в баззу данных из "Data.xlsx" документа.\n'))

wb = load_workbook(filename = 'Data.xlsx')

for i in range (2, number_of_users+2): #первая строка в документе сверочная (указаны данные, которые надо ввести, поэтому начало со второй)

    email = str(wb.active.cell(row=i, column=1).value)
    name = str(wb.active.cell(row=i, column=2).value)
    password = str(wb.active.cell(row=i, column=3).value)
    id = int(wb.active.cell(row=i, column=4).value)
    course = int(wb.active.cell(row=i, column=5).value)
    group = str(wb.active.cell(row=i, column=6).value)
    faculty = str(wb.active.cell(row=i, column=7).value)
    role = str(wb.active.cell(row=i, column=8).value)

    #print(email, "\t", name, "\t", password, "\t", id, "\t", course, "\t", group, "\t", faculty, "\t", role) *проверочная строка выводая*


    new_user = User(email=email, name=name, password=generate_password_hash(password, method='sha256'),
    id=id, course=course, group=group, role=role, faculty=faculty)

    db.session.add(new_user)
    db.session.commit()

print('Программа завершила свою работу успешно. Проверьте базу данных "db.sqlite"')